#!/usr/bin/env python3
"""
Convert all xlsx files in data/2026/ to CSV files matching the question_bank.csv format.
CSV format: "type","question","options","answer"
Options are separated by real newlines: "A. text\nB. text\nC. text\nD. text"
"""

import os
import re
import csv
import openpyxl

XLSX_DIR = os.path.join(os.path.dirname(__file__), '2026')
CSV_DIR = os.path.join(os.path.dirname(__file__), '2026')


def clean(val):
    """Strip a cell value of leading/trailing whitespace and normalize internal whitespace."""
    if val is None:
        return ''
    s = str(val)
    # Remove leading/trailing whitespace (including newlines)
    s = s.strip()
    return s


def normalize_options(raw):
    """
    Normalize options field to the canonical format:
      A. text\nB. text\nC. text\nD. text
    Handles:
      - Real newlines between options
      - Literal \\n between options
      - Space-only separators (A. text B. text C. text D. text)
    """
    if not raw:
        return ''

    # Replace literal \n (two chars) with real newline
    s = raw.replace('\\n', '\n')

    # Insert a real newline before every option marker B-G (handles space-only separation)
    # This covers cases like "A. text B. text C. text D. text"
    s = re.sub(r'\s+(?=[B-G]\.\s)', '\n', s)

    # Now split on newlines, clean each line, rejoin
    lines = [line.strip() for line in s.split('\n')]
    lines = [l for l in lines if l]  # remove empty lines

    return '\n'.join(lines)


def normalize_answer(raw):
    """Clean answer field."""
    if not raw:
        return ''
    s = str(raw).strip()
    # Remove surrounding quotes, extra whitespace/newlines
    s = re.sub(r'[\n\r]+', '', s)
    s = s.strip()
    return s


def xlsx_to_csv(xlsx_path, csv_path):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Find the right columns by scanning header row
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    headers_clean = [clean(str(h)) if h else '' for h in headers]

    # Detect column indices (1-based) for: type, question, options, answer
    type_col = q_col = opt_col = ans_col = None
    for i, h in enumerate(headers_clean, 1):
        if '题型' in h:
            type_col = i
        elif '答案' in h:
            ans_col = i
        elif '选项' in h:
            opt_col = i
        elif '题干' in h or '题目' in h or ('序号' not in h and h and type_col is not None and opt_col is None):
            q_col = i

    # Fallback: assume columns 2=type, 3=question, 4=options, 5=answer
    if type_col is None: type_col = 2
    if q_col is None: q_col = 3
    if opt_col is None: opt_col = 4
    if ans_col is None: ans_col = 5

    rows_written = 0
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, quoting=csv.QUOTE_ALL)
        writer.writerow(['type', 'question', 'options', 'answer'])

        for row in ws.iter_rows(min_row=2, values_only=True):
            qtype = clean(row[type_col - 1])
            question = clean(row[q_col - 1])
            options_raw = str(row[opt_col - 1]) if row[opt_col - 1] is not None else ''
            answer = normalize_answer(row[ans_col - 1])

            if not qtype or not question:
                continue

            options = normalize_options(options_raw)

            writer.writerow([qtype, question, options, answer])
            rows_written += 1

    return rows_written


def main():
    xlsx_files = sorted([f for f in os.listdir(XLSX_DIR) if f.endswith('.xlsx')])
    print(f"Found {len(xlsx_files)} xlsx files\n")

    total = 0
    for fname in xlsx_files:
        xlsx_path = os.path.join(XLSX_DIR, fname)
        csv_name = os.path.splitext(fname)[0] + '.csv'
        csv_path = os.path.join(CSV_DIR, csv_name)

        try:
            count = xlsx_to_csv(xlsx_path, csv_path)
            print(f"✓ {fname}  →  {count} questions")
            total += count
        except Exception as e:
            print(f"✗ {fname}  ERROR: {e}")

    print(f"\nTotal: {total} questions across {len(xlsx_files)} files")


if __name__ == '__main__':
    main()
